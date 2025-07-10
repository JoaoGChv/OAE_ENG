from __future__ import annotations # Permite o uso de anotações de tipo de futuro (ex: list[str])
import datetime # Módulo para trabalhar com datas e horas
import os # Módulo para interagir com o sistema operacional (caminhos de arquivo, etc.)
import re # Módulo para expressões regulares
import string # Módulo para operações com strings (usado para letras do alfabeto)
from pathlib import Path # Módulo para manipulação de caminhos de arquivo de forma orientada a objetos
from typing import Dict, List, Sequence, Tuple # Tipos para anotações de tipo
import urllib.parse as _up # Importa o módulo urllib.parse como _up (para uso futuro, não usado diretamente para percent-encoding aqui)
from urllib.parse import quote as _q # Importa a função quote para percent-encoding de caracteres (usado em _to_uri_folder)
from openpyxl import Workbook, load_workbook # Classes para criar e carregar pastas de trabalho Excel
from openpyxl.styles import Alignment, Font, PatternFill # Classes para estilos de células (alinhamento, fonte, preenchimento)
from openpyxl.utils import get_column_letter # Função para converter número de coluna em letra (ex: 1 -> A)
import logging # Módulo para registro de eventos (logs)

logger = logging.getLogger(__name__) # Configura um logger para este módulo

# ---------------------------------------------------------------------------
# Estilo e constantes
# ---------------------------------------------------------------------------
COR_REVISADO = "00A8FF" # Código hexadecimal da cor para arquivos revisados (azul claro)
COR_NOVO = "91EF93" # Código hexadecimal da cor para arquivos novos (verde claro)
COR_MODIFICADO = "FFA500" # Código hexadecimal da cor para arquivos modificados (laranja)
COR_INALTERADO = "FFFFFF" # Código hexadecimal da cor para arquivos inalterados (branco)
COR_TITULO = "5B9BD5" # Código hexadecimal da cor para títulos de cabeçalho (azul escuro)

STATUS_COR = {
    "revisado": COR_REVISADO, # Mapeia status 'revisado' para sua cor
    "novo": COR_NOVO, # Mapeia status 'novo' para sua cor
    "modificado": COR_MODIFICADO, # Mapeia status 'modificado' para sua cor
    "inalterado": COR_INALTERADO, # Mapeia status 'inalterado' para sua cor
}

FONT_TITULO = Font(bold=True, size=14) # Objeto Font para títulos (negrito, tamanho 14)
ALIGN_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True) # Objeto Alignment para centralizar texto e quebrar linha

LARG_A_I = 8.43 # Largura padrão para as colunas de A a I
LARG_J = 8.43 # Largura padrão para a coluna J
LARG_GRUPO = 28.71 # Largura para a coluna de grupo de extensão
LARG_EXT = 23 # Largura para a coluna de extensão
LARG_ENT = 68 # Largura para TODAS as colunas de entrega

GRUPOS_EXT: Dict[str, Sequence[str]] = {
    "DWG/DXF": [".dwg", ".dxf"], # Grupo de extensões para arquivos CAD
    "PDF": [".pdf"], # Grupo para arquivos PDF
    "XLS/XLSX": [".xls", ".xlsx"], # Grupo para arquivos Excel
    "DOC/DOCX": [".doc", ".docx"], # Grupo para arquivos Word
}

REV_REGEX = re.compile(r"([-._])R(\d{1,3})$", re.IGNORECASE) # Expressão regular para encontrar padrões de revisão (ex: -R01, .R123)

# ---------------------------------------------------------------------------
# Utilidades de coluna/planilha
# ---------------------------------------------------------------------------
def _col_idx(letra: str) -> int:
    # Converte uma letra de coluna (ex: "A", "J") para seu índice numérico (1, 10).
    return string.ascii_uppercase.index(letra.upper()) + 1 # Retorna o índice baseado em 1

def _merge_ai(ws, row: int, value: str = "", font: Font | None = None):
    # Mescla as células de A a I em uma linha específica, define um valor, alinhamento e fonte.
    ws.merge_cells(start_row=row, start_column=_col_idx("A"),
                   end_row=row, end_column=_col_idx("I")) # Mescla as células
    c = ws.cell(row=row, column=_col_idx("A"), value=value) # Obtém a célula mesclada
    c.alignment = ALIGN_CENTRO # Aplica alinhamento central
    if font:
        c.font = font # Aplica a fonte se fornecida

def _set_widths(ws):
    # Define as larguras das colunas A a I, J, K e L na planilha.
    for col in range(_col_idx("A"), _col_idx("I") + 1):
        ws.column_dimensions[get_column_letter(col)].width = LARG_A_I # Define largura para colunas A-I
    ws.column_dimensions["J"].width = LARG_J # Define largura para coluna J
    ws.column_dimensions["K"].width = LARG_GRUPO # Define largura para coluna K (Grupo)
    ws.column_dimensions["L"].width = LARG_EXT # Define largura para coluna L (Extensão)

# ---------------------------------------------------------------------------
# Chave, revisão, grupo
# ---------------------------------------------------------------------------
def _key(nome: str) -> Tuple[str, str]:
    # Extrai o nome base normalizado (sem revisão e com hífens no lugar de underscores) e a extensão de um nome de arquivo.
    base, ext = os.path.splitext(nome) # Separa nome base e extensão
    base_norm = base.replace("_", "-") # Normaliza underscores para hífens
    m = REV_REGEX.search(base_norm) # Procura por padrão de revisão
    if m:
        base_norm = base_norm[: m.start()] # Remove a revisão do nome base
    return base_norm.lower(), ext.lower() # Retorna nome base e extensão em minúsculas

def _extrair_rev(nome: str) -> str:
    # Extrai a revisão de um nome de arquivo, formatando-a como "RXX" (ex: "R01").
    m = REV_REGEX.search(os.path.splitext(nome)[0]) # Procura por padrão de revisão no nome sem extensão
    return f"R{int(m.group(2)):02d}" if m else "" # Formata e retorna a revisão, ou string vazia se não encontrada

def _classificar_extensao(ext: str) -> str:
    # Classifica a extensão de um arquivo em um grupo predefinido (ex: "DWG/DXF", "PDF").
    for grp, exts in GRUPOS_EXT.items(): # Itera sobre os grupos de extensão definidos
        if ext.lower() in exts: # Verifica se a extensão (em minúsculas) está no grupo
            return grp # Retorna o nome do grupo
    return "Outros" # Retorna "Outros" se a extensão não se encaixar em nenhum grupo

# ---------------------------------------------------------------------------
# NOVO helper: monta URI file:/// sem percent-encoding
# ---------------------------------------------------------------------------
from urllib.parse import quote as _q # Reimporta quote explicitamente para este bloco (já importado acima)

def _to_uri_folder(path: str) -> str:
    # Converte um caminho de sistema de arquivos em uma URI no formato 'file:///.../?open' para hyperlinks no Excel.
    path = path.replace("\\", "/").rstrip("/") # Normaliza barras e remove barra final se existir
    path = re.sub(r"^([A-Za-z]):/{2,}", r"\1:/", path) # Corrige paths como "G://Folder" para "G:/Folder"
    safe = "/:-_." # Caracteres considerados "seguros" que não precisam de percent-encoding
    encoded = "".join(
        ch if (ch.isalnum() or ch in safe) else _up.quote(ch, safe="") # Codifica caracteres não seguros, mas mantendo os definidos em 'safe'
        for ch in path
    )
    # Adiciona barra final e "#" para evitar que o Excel modifique o hyperlink e "?open" para abrir a pasta
    return "file:///" + encoded + "/?open"

# ---------------------------------------------------------------------------
# Helper que hidrata links faltantes em colunas antigas
# ---------------------------------------------------------------------------
def _hidratar_hyperlinks(ws, linha_titulo: int, dir_base: str) -> None:
    """Preenche hyperlinks de colunas antigas apontando para as pastas corretas."""
    col_first = _col_idx("M") # Coluna inicial onde as entregas começam (M = 13)
    mapas: Dict[int, str] = {} # Dicionário para mapear coluna para o caminho da pasta de entrega

    for col in range(col_first, ws.max_column + 1): # Itera sobre todas as colunas de entrega
        cab = ws.cell(row=linha_titulo, column=col).value or "" # Obtém o cabeçalho da coluna (ex: "1.AP - Entrega-1")
        m = re.search(r"(1\.AP|2\.PE)\s*-\s*Entrega-\s*(\d+)", str(cab)) # Procura por padrão de nome de entrega
        if m:
            numero = m.group(2) # Extrai o número da entrega
            subdir = "AP" if m.group(1).startswith("1") else "PE" # Determina o subdiretório (AP ou PE)
            prefixo = f"{m.group(1)} - Entrega-" # Reconstrói o prefixo (ex: "1.AP - Entrega-")
            pasta = os.path.join(dir_base, subdir, f"{prefixo}{numero}") # Monta o caminho esperado da pasta de entrega

            if not os.path.exists(pasta):
                # Se a pasta não existir no local padrão, procura por uma versão "OBSOLETO"
                pai = os.path.dirname(pasta) # Diretório pai (AP ou PE)
                padrao = f"{prefixo}{numero}-OBSOLETO" # Padrão para pastas obsoletas
                # Lista diretórios no pai que começam com o padrão OBSOLETO
                candidatos = [p for p in os.listdir(pai) if p.startswith(padrao)] if os.path.isdir(pai) else []
                if candidatos:
                    pasta = os.path.join(pai, candidatos[0]) # Usa a primeira pasta obsoleta encontrada

            mapas[col] = pasta # Mapeia a coluna para o caminho da pasta de entrega

    for row in range(linha_titulo + 1, ws.max_row + 1): # Itera sobre todas as linhas de dados na planilha
        for col, pasta in mapas.items(): # Itera sobre as colunas que representam entregas
            cel = ws.cell(row=row, column=col) # Obtém a célula atual
            if cel.value: # Se a célula tiver um valor (nome de arquivo)
                link = _to_uri_folder(pasta) # Converte o caminho da pasta para URI de hyperlink
                logger.debug("[hidratar] row %s col %s → %s", row, col, link) # Loga a criação do hyperlink
                cel.hyperlink = link # Define o hyperlink para a célula

# ---------------------------------------------------------------------------
# Função principal – lógica original preservada
# ---------------------------------------------------------------------------
def criar_ou_atualizar_planilha(
    caminho_excel: str | Path, # Caminho completo do arquivo Excel a ser criado/atualizado
    tipo_entrega: str, # Tipo da entrega (ex: "AP", "PE")
    num_entrega: int, # Número sequencial da entrega
    diretorio_base: str, # Diretório raiz onde as entregas são organizadas
    arquivos: List[Tuple[str, str, int, str, str]], # Lista de arquivos com suas informações (revisão, nome, tamanho, caminho completo, data de modificação)
    estado_anterior: Dict[str, Dict[str, object]] | None = None, # Dicionário com o estado dos arquivos da execução anterior
):
    # Converte o caminho do Excel para um objeto Path
    caminho_excel = Path(caminho_excel)
    # Caso queria voltar ao que era originalmente, comente ou apague a variável "criado" logo na linha abaixo:
    wb, ws, criado = _abrir_ou_criar_wb(caminho_excel, diretorio_base) # Abre um workbook existente ou cria um novo

    ws.freeze_panes = "J10" # Congela os painéis para fixar as linhas e colunas de cabeçalho
    _set_widths(ws) # Define as larguras das colunas

    linha_titulo = 9 # Linha onde começa o cabeçalho das entregas
    col_ent = _col_idx("M") # Coluna inicial para a nova entrega (M)
    ws.insert_cols(col_ent) # Insere uma nova coluna na posição 'col_ent' para a entrega atual
    col_prev = col_ent + 1 # Coluna anterior (para compatibilidade, pode ser desnecessária dependendo da lógica)

    for c in range(col_ent, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = LARG_ENT # Define largura para todas as colunas de entrega

    _pintar_cabecalho_titulos(ws, linha_titulo) # Pinta as células de cabeçalho (Grupo, Extensão)
    prefixos = {"AP": "1.AP - Entrega-", "PE": "2.PE - Entrega-"} # Define prefixos para nomes de entrega
    # Cria o cabeçalho para a nova coluna de entrega (ex: "1.AP - Entrega- X")
    cab = ws.cell(row=linha_titulo, column=col_ent,
                  value=f"{prefixos.get(tipo_entrega,'ENT')} {num_entrega}")
    cab.alignment = ALIGN_CENTRO # Centraliza o texto do cabeçalho
    cab.font = Font(bold=True) # Define a fonte como negrito
    cab.fill = PatternFill(start_color=COR_TITULO,
                           end_color=COR_TITULO,
                           fill_type="solid") # Preenche a célula com a cor de título

    # Caso queria voltar ao que era originalmente, descomente a linha abaixo:
    # snapshot_ant = _carregar_snapshot(ws, linha_titulo, col_prev, estado_anterior)

    # Seguindo: E comente as linhas do IF até snapshot_ant (...).
    # Lógica para carregar o snapshot: Se a planilha foi criada agora E há estado anterior,
    # reconstrói o snapshot a partir do estado_anterior. Caso contrário, carrega da planilha.
    if criado and estado_anterior: # Verifica se a planilha foi recém-criada e há dados de estado anterior
        snapshot_ant: Dict[Tuple[str, str], Dict[str, object]] = {} # Inicializa um snapshot vazio
        linha_tmp = ws.max_row + 1 # Começa a adicionar informações de arquivos anteriores a partir da próxima linha disponível
        for chave, dados in estado_anterior.items(): # Itera sobre os arquivos do estado anterior
            try:
                base, ext = chave.split("|", 1) # Divide a chave (nome_base|extensao)
            except ValueError:
                continue # Pula se a chave não estiver no formato esperado
            base = base.lower() # Normaliza nome base para minúsculas
            ext = ext.lower() # Normaliza extensão para minúsculas
            # Adiciona informações de grupo e extensão nas colunas K e L para esses arquivos (simulando que eles existiam)
            ws.cell(row=linha_tmp, column=_col_idx("K"),
                    value=_classificar_extensao(ext))
            ws.cell(row=linha_tmp, column=_col_idx("L"), value=ext)
            # Adiciona as informações do arquivo ao snapshot_ant (com base no estado anterior)
            snapshot_ant[(base, ext)] = {
                "rev": dados.get("revisao", ""), # Revisão do estado anterior
                "tam": dados.get("tamanho"), # Tamanho do estado anterior
                "ts": dados.get("timestamp"), # Timestamp do estado anterior
                "row": linha_tmp, # Linha onde esta informação seria registrada na planilha
            }
            linha_tmp += 1 # Avança para a próxima linha temporária
    else:
        # Se a planilha já existia ou não há estado anterior, carrega o snapshot da própria planilha
        snapshot_ant = _carregar_snapshot(ws, linha_titulo, col_prev, estado_anterior)

    atual_info = {} # Dicionário para armazenar as informações dos arquivos atuais
    for rev, nome, tam, full_path, _ in arquivos: # Itera sobre a lista de arquivos da entrega atual
        base, ext = _key(nome) # Obtém a chave (nome base normalizado e extensão)
        # Popula atual_info com dados do arquivo atual
        atual_info[(base, ext)] = {
            "nome": nome, # Nome completo do arquivo
            "rev": rev or _extrair_rev(nome), # Revisão (usa a fornecida ou extrai do nome)
            "tam": tam, # Tamanho do arquivo
            "ts": os.path.getmtime(full_path), # Timestamp da última modificação
            "dir": os.path.dirname(full_path), # Diretório onde o arquivo está localizado
            "grupo": _classificar_extensao(ext), # Grupo de extensão do arquivo
            "ext": ext, # Extensão do arquivo
        }

    # Pinta as células da coluna da entrega anterior como "inalterado" (branco)
    for snap in snapshot_ant.values():
        ws.cell(row=snap["row"], column=col_ent).fill = PatternFill(
            start_color=COR_INALTERADO, end_color=COR_INALTERADO, fill_type="solid"
        )

    linha_cursor = ws.max_row + 1 # Cursor para a próxima linha disponível para novos arquivos

    for key_, info in atual_info.items(): # Itera sobre os arquivos da entrega atual
        snap = snapshot_ant.get(key_) # Tenta encontrar o arquivo no snapshot anterior
        if snap:
            row = snap["row"] # Se encontrado, usa a linha existente
            status = _determinar_status(info, snap) # Determina o status (revisado, modificado, inalterado)
        else:
            row = linha_cursor # Se não encontrado, é um arquivo novo, usa a próxima linha
            linha_cursor += 1 # Avança o cursor para a próxima linha
            # Preenche as colunas de grupo e extensão para o novo arquivo
            ws.cell(row=row, column=_col_idx("K"), value=info["grupo"])
            ws.cell(row=row, column=_col_idx("L"), value=info["ext"])
            status = "novo" # Define o status como "novo"

        cor = STATUS_COR[status] # Obtém a cor correspondente ao status
        cel = ws.cell(row=row, column=col_ent, value=info["nome"]) # Define o nome do arquivo na célula da entrega atual
        cel.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid") # Pinta a célula com a cor do status

        # ---------- hyperlink (para a PASTA, barra final garantida) ----------
        folder = info["dir"] # Diretório do arquivo
        link = _to_uri_folder(folder) # Converte o diretório para URI de hyperlink
        logger.debug("[nova-col] linha %s | %s", row, link) # Loga a criação do hyperlink
        cel.hyperlink = link # Atribui o hyperlink à célula
    # ---- hidrata colunas antigas ------------------------------------------
    _hidratar_hyperlinks(ws, linha_titulo, diretorio_base) # Preenche hyperlinks em colunas de entregas anteriores
    wb.save(caminho_excel) # Salva o workbook Excel
    logger.info("Planilha atualizada: %s", caminho_excel) # Loga que a planilha foi atualizada

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _determinar_status(atual: dict, snap: dict | None) -> str:
    # Determina o status de um arquivo (novo, revisado, modificado, inalterado) comparando informações atuais com o snapshot anterior.
    if snap is None:
        return "novo" # Se não há snapshot anterior, é novo
    if atual["rev"] != snap.get("rev", ""):
        return "revisado" # Se a revisão mudou, é revisado
    if snap.get("tam") is not None and atual["tam"] != snap["tam"]:
        return "modificado" # Se o tamanho mudou e o tamanho anterior existe, é modificado
    if snap.get("ts") is not None and atual["ts"] != snap["ts"]:
        return "modificado" # Se o timestamp mudou e o timestamp anterior existe, é modificado
    return "inalterado" # Caso contrário, é inalterado

def _carregar_snapshot(ws, linha_titulo: int, col_prev: int,
                       estado_anterior: Dict[str, Dict[str, object]] | None = None):
    # Carrega um "snapshot" dos arquivos da planilha, ou a partir de um estado anterior JSON se disponível.
    snap: Dict[Tuple[str, str], Dict[str, object]] = {} # Dicionário para armazenar o snapshot
    for row in range(linha_titulo + 1, ws.max_row + 1): # Itera sobre as linhas de dados da planilha
        nome = ws.cell(row=row, column=col_prev).value # Obtém o nome do arquivo da coluna 'col_prev'
        if not nome:
            continue # Pula se não houver nome
        base, ext = _key(nome) # Obtém a chave do arquivo
        rev_plan = _extrair_rev(nome) # Extrai a revisão do nome do arquivo na planilha
        # Tenta obter dados do arquivo do estado anterior (JSON) se fornecido
        dados_json = estado_anterior.get(f"{base}|{ext}", {}) if estado_anterior else {}
        # Popula o snapshot com informações do arquivo (preferindo o JSON para revisão, tamanho, timestamp)
        snap[(base, ext)] = {
            "rev": dados_json.get("revisao", rev_plan), # Usa revisão do JSON se existir, senão da planilha
            "tam": dados_json.get("tamanho"), # Usa tamanho do JSON se existir
            "ts": dados_json.get("timestamp"), # Usa timestamp do JSON se existir
            "row": row, # Registra a linha do arquivo na planilha
        }
    return snap # Retorna o snapshot carregado

# ---------------------------------------------------------------------------
# Layout inicial / cabeçalhos
# ---------------------------------------------------------------------------
def _abrir_ou_criar_wb(caminho: Path, dir_base: str):
    # Abre uma pasta de trabalho Excel existente ou cria uma nova.
    # Retorna o workbook, a worksheet ativa e um booleano indicando se foi criada (True) ou aberta (False).
    if caminho.exists(): # Verifica se o arquivo Excel já existe
        wb = load_workbook(caminho) # Carrega o workbook existente
        return wb, wb.active, False # Retorna o workbook, a planilha ativa e False (não foi criada)
    wb = Workbook() # Cria um novo workbook
    ws = wb.active # Obtém a planilha ativa (primeira)
    ws.title = "GRD" # Define o título da planilha como "GRD"
    _montar_cabecalho(ws, dir_base) # Monta o cabeçalho inicial da planilha
    return wb, ws, True # Retorna o novo workbook, a planilha ativa e True (foi criada)

def _montar_cabecalho(ws, dir_base: str):
    # Monta o cabeçalho inicial da planilha com informações do projeto e legendas de cores.
    hoje = datetime.datetime.now().strftime("%d-%m-%Y_%H-%M") # Formata a data e hora atuais
    _merge_ai(ws, 1, "OLIVEIRA ARAÚJO ENGENHARIA", FONT_TITULO) # Título principal
    _merge_ai(ws, 2, "Lista de arquivos de projetos entregues com controle de revisões") # Subtítulo
    _merge_ai(ws, 3, f"Diretório: {dir_base}") # Informa o diretório base
    _merge_ai(ws, 4, f"Data de emissão: {hoje}") # Informa a data de emissão

    legendas = [
        ("Arquivo Revisado", COR_REVISADO), # Legenda para arquivo revisado
        ("Arquivo Novo", COR_NOVO), # Legenda para arquivo novo
        ("Arquivo Modificado s/ Atualizar Revisão", COR_MODIFICADO), # Legenda para arquivo modificado sem atualização de revisão
        ("Arquivo Inalterado", COR_INALTERADO), # Legenda para arquivo inalterado
    ]
    for i, (txt, cor) in enumerate(legendas, start=5): # Itera sobre as legendas a partir da linha 5
        _merge_ai(ws, i, txt) # Mescla células e define o texto da legenda
        c = ws.cell(row=i, column=_col_idx("A")) # Obtém a célula para aplicar o estilo
        c.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid") # Preenche a célula com a cor da legenda
        c.font = Font(bold=True) # Define a fonte da legenda como negrito

def _pintar_cabecalho_titulos(ws, linha: int):
    # Pinta e formata as células de cabeçalho para as colunas "Grupo" (K) e "Extensão" (L).
    # pinta K (Grupo) e L (Extensão)
    for col in (_col_idx("K"), _col_idx("L")): # Itera sobre as colunas K e L
        cel = ws.cell(row=linha, column=col) # Obtém a célula do cabeçalho
        cel.fill = PatternFill(start_color=COR_TITULO,
                               end_color=COR_TITULO,
                               fill_type="solid") # Pinta a célula com a cor de título
        cel.alignment = ALIGN_CENTRO # Centraliza o texto
        cel.font = Font(bold=True) # Define a fonte como negrito