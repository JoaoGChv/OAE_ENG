from __future__ import annotations
import re
import os
import sys
import json
import shutil
import hashlib
import logging
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict

# --------------------- CONFIGURAÇÕES ---------------------
JSON_CONTADORES_DIR = r"G:\Drives compartilhados\OAE - SCRIPTS\SCRIPTS\tmp_joaoG\JSON_tmp_joao"
PROJETOS_JSON = r"G:\Drives compartilhados\OAE-JSONS\diretorios_projetos.json"
SCRIPT_DIR = Path(__file__).parent
NOMENCLATURA_REGRAS_JSON = r"G:\Drives compartilhados\OAE - SCRIPTS\SCRIPTS\tmp_joaoG\Melhorias\Código_reformulado_teste\OAE_ENG\nomenclaturas.json"
ULTIMO_DIRETORIO_JSON = "ultimo_diretorio.json"
HISTORICO_JSON = "historico_arquivos.json"
JSON_FILE_PATH = "dados_projetos.json"
MARGIN_SIZE = 10
TEMPLATE_XLSX = Path(__file__).with_name("GRD_template.xlsx")
print("DEBUG-PATH:", TEMPLATE_XLSX)


LOG_FILENAME = "debug_entregas.log"
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%d/%m/%Y %H:%M:%S",
    handlers=[
        logging.FileHandler(LOG_FILENAME, encoding="utf-8"),
        logging.StreamHandler(sys.stdout)
    ]
)

AP_PREFIX = "1.AP - Entrega-"
PE_PREFIX = "2.PE - Entrega-"
ENTREGA_RE = re.compile(r"^(1\.AP|2\.PE) - Entrega-(\d+)$")


# -----------------------------------------------------
# FUNÇÕES AUXILIARES ORIGINAIS
# -----------------------------------------------------
def _listar_entregas_tipo(pasta: Path, prefixo: str) -> list[Path]:
    return sorted(
        [p for p in pasta.iterdir()
         if p.is_dir() and p.name.startswith(prefixo) and not p.name.endswith("-OBSOLETO")],
        key=lambda p: int(ENTREGA_RE.match(p.name).group(2))
    )

def _proximo_num_entrega(pasta_entregas: Path, prefixo: str) -> int:
    ativas = _listar_entregas_tipo(pasta_entregas, prefixo)
    if not ativas:
        return 1
    ultimo = ENTREGA_RE.match(ativas[-1].name)
    return int(ultimo.group(2)) + 1

def _marcar_obsoleta(p: Path):
    destino = p.with_name(p.name + "-OBSOLETO")
    seq = 1
    while destino.exists():
        seq += 1
        destino = p.with_name(p.name + f"-OBSOLETO{seq}")
    p.rename(destino)
    logging.info("Renomeada %s ➜ %s", p.name, destino.name)

def _hash_file(path: Path, buf=8192) -> str:
    h = hashlib.md5()
    with path.open("rb") as f:
        while chunk := f.read(buf):
            h.update(chunk)
    return h.hexdigest()

def obter_entrega_anterior(pasta_entregas: Path) -> Optional[Path]:
    entregas = sorted(
        [p for p in pasta_entregas.iterdir()
         if p.is_dir()
         and p.name.startswith("Entrega_")
         and p.name.split("_")[1][:2].isdigit()
         and not p.name.endswith("_OBS")],
        key=lambda p: int(p.name.split("_")[1][:2])
    )
    return entregas[-1] if entregas else None

def listar_arquivos_entrega(pasta: Path) -> list[Path]:
    return [p for p in pasta.iterdir() if p.is_file()]

def comparar_arquivos(pasta_nova: Path, pasta_ant: Optional[Path]) -> dict:
    atual     = {p.name: p for p in listar_arquivos_entrega(pasta_nova)}
    anterior  = {p.name: p for p in listar_arquivos_entrega(pasta_ant)} if pasta_ant else {}
    resultado: Dict[str, dict] = {}

    for nome, p in atual.items():
        if nome == "_controle_entrega.json":
            continue
        if nome in anterior:
            ig = _hash_file(p) == _hash_file(anterior[nome])
            resultado[nome] = {
                "status": "nao_modificado" if ig else "modificado",
                "versao_anterior": str(anterior[nome])
            }
        else:
            resultado[nome] = {"status": "novo"}

    for nome, p_old in anterior.items():
        if nome not in resultado and nome != "_controle_entrega.json":
            resultado[nome] = {"status": "removido", "versao_anterior": str(p_old)}
    return resultado

def gerar_arquivo_controle(nova_pasta: Path, comparacao: dict):
    with (nova_pasta / "_controle_entrega.json").open("w",  encoding="utf-8") as f:
        json.dump(comparacao, f, indent=4, ensure_ascii=False)

def salvar_historico_global_entregas(pasta_entregas: Path, registro: dict):
    historico_path = pasta_entregas / "historico_entregas.json"
    historico = []
    if historico_path.exists():
        with open(historico_path, "r", encoding="utf-8") as f:
            try:
                historico = json.load(f)
            except Exception:
                historico = []
    historico.append(registro)
    with open(historico_path, "w", encoding="utf-8") as f:
        json.dump(historico, f, indent=4, ensure_ascii=False)

def processar_entrega_arquivos_tipo(arquivos: list[Path], pasta_entregas: Path, tipo: str) -> Path:
    tipo_subpasta = 'AP' if tipo == "AP" else 'PE'
    pasta_tipo = pasta_entregas / tipo_subpasta
    pasta_tipo.mkdir(exist_ok=True, parents=True)

    prefixo = AP_PREFIX if tipo == "AP" else PE_PREFIX
    etapa = 1 if tipo == "AP" else 2

    ativas = _listar_entregas_tipo(pasta_tipo, prefixo)
    entrega_ativa = ativas[-1] if ativas else None

    n     = _proximo_num_entrega(pasta_tipo, prefixo)
    nova  = pasta_tipo / f"{prefixo}{n}"
    nova.mkdir(parents=True, exist_ok=False)
    logging.debug("Criada nova entrega: %s", nova)

    comp = comparar_arquivos(nova, entrega_ativa)

    if entrega_ativa:
        _marcar_obsoleta(entrega_ativa)

    for src in arquivos:
        shutil.copy2(src, nova / src.name)

    comp.update({"tipo_entrega": tipo, "etapa": etapa})
    with (nova / "_controle_entrega.json").open("w", encoding="utf-8") as f:
        json.dump(comp, f, indent=4, ensure_ascii=False)

    registro_historico = {
        "data": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "tipo_entrega": tipo,
        "etapa": etapa,
        "pasta_entrega": str(nova),
        "arquivos_entregues": [src.name for src in arquivos],
    }
    salvar_historico_global_entregas(pasta_entregas, registro_historico)

    try:
        criar_arquivo_controle(pasta_entregas)
        logging.debug("GRD.xlsx atualizado em %s", pasta_entregas)
    except Exception:
        logging.exception("Falha ao gerar GRD.xlsx")

    return nova

def carregar_regras_nomenclatura(projeto_num: str) -> dict:
    """
    Lê o JSON completo e retorna apenas o dicionário de 'campos' para o projeto.
    Se não existir, retorna {}.
    """
    if not os.path.exists(NOMENCLATURA_REGRAS_JSON):
        logging.warning("Arquivo %s não encontrado", NOMENCLATURA_REGRAS_JSON)
        return {}
    try:
        with open(NOMENCLATURA_REGRAS_JSON, encoding="utf-8") as f:
            todas_regras = json.load(f)
    except json.JSONDecodeError as e:
        logging.error("JSON inválido em %s – %s", NOMENCLATURA_REGRAS_JSON, e)
        return {}

    projeto_key = str(projeto_num)
    projeto_entry = todas_regras.get(projeto_key)
    if not projeto_entry or "campos" not in projeto_entry:
        logging.warning("Nenhuma regra encontrada para o projeto %s", projeto_key)
        return {}
    return projeto_entry  # será um dict com "campos": [...] e possivelmente "REVISÃO_ESPECIAL"

def caminho_contador(projeto_num: str) -> str:
    os.makedirs(JSON_CONTADORES_DIR, exist_ok=True)
    return os.path.join(JSON_CONTADORES_DIR, f"contador_entregas_{projeto_num}.json")

def obter_proximo_indice(projeto_num: str) -> int:
    fp = caminho_contador(projeto_num)
    data = carregar_json(fp) or {"proximo": 1}
    return data["proximo"]

def incrementar_indice(projeto_num: str):
    fp = caminho_contador(projeto_num)
    data = carregar_json(fp) or {"proximo": 1}
    data["proximo"] += 1
    salvar_json(fp, data)

def carregar_historico_entregas(projeto_num: str) -> dict:
    fp = caminho_contador(projeto_num)
    if os.path.exists(fp):
        with open(fp, "r", encoding="utf-8") as f:
            dados = json.load(f)
    else:
        dados = {}
    dados.setdefault("proximo", 1)
    dados.setdefault("entregas", [])
    return dados

def salvar_historico_entregas(projeto_num: str, data: dict) -> None:
    fp = caminho_contador(projeto_num)
    try:
        with open(fp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception as err:
        messagebox.showerror("Erro", f"Falha ao salvar histórico de entregas:\n{err}")
        raise SystemExit

def carregar_projetos():
    if not os.path.exists(PROJETOS_JSON):
        messagebox.showerror("Erro", f"Arquivo não encontrado: {PROJETOS_JSON}")
        return {}
    try:
        with open(PROJETOS_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        messagebox.showerror("Erro", f"Erro ao decodificar o JSON: {e}")
        return {}

def carregar_ultimo_diretorio():
    if os.path.exists(ULTIMO_DIRETORIO_JSON):
        with open(ULTIMO_DIRETORIO_JSON, "r", encoding="utf-8") as f:
            return json.load(f).get("ultimo_diretorio", os.getcwd())
    return os.getcwd()

def salvar_ultimo_diretorio(d):
    with open(ULTIMO_DIRETORIO_JSON, "w", encoding="utf-8") as f:
        json.dump({"ultimo_diretorio": d}, f)

def atualizar_historico(lista_arquivos, c=HISTORICO_JSON):
    h = {}
    if os.path.exists(c):
        with open(c, "r", encoding="utf-8") as f:
            h = json.load(f)
    for arq in lista_arquivos:
        dm = os.path.getmtime(arq)
        if arq not in h or h[arq]["data"] != dm:
            h[arq] = {"numero": len(h)+1, "data": dm, "status": "Atual"}
    mr = max(h, key=lambda x: h[x]["data"])
    for a in h:
        h[a]["status"] = "Atual" if a == mr else "Obsoleto"
    with open(c, "w", encoding="utf-8") as f:
        json.dump(h, f, indent=4, ensure_ascii=False)
    return h

def pos_processamento(*args):
    messagebox.showinfo("Concluído", "Processo concluído com sucesso.")
    sys.exit(0)


# -----------------------------------------------------
# FUNÇÕES DE TOKENIZAÇÃO E VALIDAÇÃO 
# -----------------------------------------------------
def split_including_separators(nome_sem_ext: str, nomenclatura: dict) -> list[str]:
    tokens: list[str] = []
    i = 0
    while i < len(nome_sem_ext):
        c = nome_sem_ext[i]
        if c in ['-', '.']:
            tokens.append(c)
            i += 1
        else:
            j = i
            while j < len(nome_sem_ext) and nome_sem_ext[j] not in ['-', '.']:
                j += 1
            tokens.append(nome_sem_ext[i:j])
            i = j
    return tokens

def verificar_tokens(tokens: list[str], nomenclatura: dict) -> list[str]:
    if not nomenclatura or "campos" not in nomenclatura:
        return ["mismatch"] * len(tokens)

    campos_cfg = nomenclatura["campos"]
    tokens_esperados: list[tuple[str, dict | str]] = []
    for idx, cinfo in enumerate(campos_cfg):
        tokens_esperados.append(("campo", cinfo))
        if idx < len(campos_cfg) - 1:
            sep_ = cinfo.get("separador", "-")
            tokens_esperados.append(("sep", sep_))

    result_tags: list[str] = []
    idx_exp = 0
    idx_tok = 0

    while idx_tok < len(tokens) and idx_exp < len(tokens_esperados):
        t = tokens[idx_tok]
        tipo_esp, conteudo_esp = tokens_esperados[idx_exp]

        if tipo_esp == "sep":
            if t == conteudo_esp:
                result_tags.append("ok")
            else:
                result_tags.append("mismatch")
            idx_tok += 1
            idx_exp += 1
        else:
            cinfo = conteudo_esp
            tipo_campo = cinfo.get("tipo", "Fixo")
            fixos = cinfo.get("valores_fixos", [])
            if tipo_campo == "Fixo" and fixos:
                lista_val_permitido = []
                for f in fixos:
                    if isinstance(f, dict):
                        lista_val_permitido.append(f.get("value", ""))
                    else:
                        lista_val_permitido.append(str(f))
                if lista_val_permitido and t not in lista_val_permitido:
                    result_tags.append("mismatch")
                else:
                    result_tags.append("ok")
            else:
                result_tags.append("ok")
            idx_tok += 1
            idx_exp += 1

    while idx_tok < len(tokens):
        result_tags.append("mismatch")
        idx_tok += 1
    while idx_exp < len(tokens_esperados):
        result_tags.append("missing")
        idx_exp += 1

    return result_tags


# -----------------------------------------------------
# FLUXO DE JANELAS
# -----------------------------------------------------
def janela_selecao_projeto(master):
    root = master
    root.title("Selecionar Projeto")
    root.geometry("600x400")
    projetos_dict = carregar_projetos()
    if not projetos_dict:
        messagebox.showerror("Erro", "Nenhum projeto encontrado ou erro no arquivo JSON.")
        return None, None

    p_conv = []
    for num, c_full in projetos_dict.items():
        nm = os.path.basename(c_full)
        p_conv.append((num, nm, c_full))
    sel = {"numero": None, "caminho": None}

    def filtrar(*args):
        t = entrada.get().lower()
        tree.delete(*tree.get_children())
        for n, nm, co in p_conv:
            if t in nm.lower():
                tree.insert("", tk.END, values=(n, nm, co))
        all_iid = tree.get_children()
        if len(all_iid) == 1:
            tree.selection_set(all_iid[0])

    def confirmar():
        si = tree.selection()
        if not si:
            messagebox.showinfo("Info", "Selecione um projeto.")
            return
        v = tree.item(si[0], "values")
        sel["numero"] = v[0]
        sel["caminho"] = v[2]
        root.withdraw()
        Disciplinas_Detalhes_Projeto(sel["numero"], sel["caminho"], master=root)

    frame = tk.Frame(root)
    frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)
    tk.Label(frame, text="Digite nome ou parte do nome do projeto:", font=("Arial", 10)).pack(anchor="w")
    entrada = tk.Entry(frame)
    entrada.pack(fill=tk.X)
    entrada.bind("<KeyRelease>", filtrar)
    entrada.bind("<Return>", lambda e: confirmar())

    cols = ("Número", "Nome do Projeto", "Caminho Original")
    tree = ttk.Treeview(frame, columns=cols, show="headings", height=10)
    tree.heading("Número", text="Número")
    tree.heading("Nome do Projeto", text="Nome do Projeto")
    tree.heading("Caminho Original", text="Caminho Original")
    tree.column("Número", width=80)
    tree.column("Nome do Projeto", width=300)
    tree.column("Caminho Original", width=0, stretch=False, minwidth=0)
    tree.pack(fill=tk.BOTH, expand=True)

    for n, nm, co in p_conv:
        tree.insert("", tk.END, values=(n, nm, co))

    bf = tk.Frame(frame)
    bf.pack(pady=5)
    ttk.Button(bf, text="Confirmar", command=confirmar).pack(side=tk.LEFT, padx=5)
    ttk.Button(bf, text="Cancelar", command=root.destroy).pack(side=tk.LEFT, padx=5)
    root.mainloop()
    return sel["numero"], sel["caminho"]


def Disciplinas_Detalhes_Projeto(numero, caminho, master=None):
    d_path = os.path.join(caminho, "3 Desenvolvimento")
    if not os.path.exists(d_path):
        messagebox.showerror("Erro", "A pasta de disciplinas não foi encontrada.")
        return

    discip_win = tk.Toplevel(master)
    discip_win.title(f"Gerenciador de Projetos - Projeto {numero}")
    discip_win.geometry("900x600")

    hd = tk.Label(
        discip_win,
        text=f"Projeto {numero} - Selecione a disciplina para entrega",
        font=("Helvetica", 14, "bold"),
        anchor="w"
    )
    hd.pack(fill=tk.X, padx=10, pady=5)

    cols = ["Nome", "Data de Modificação", "Tipo", "Tamanho"]
    tree = ttk.Treeview(discip_win, columns=cols, show="headings", height=20)
    tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=200 if c == "Nome" else 150, anchor="w")

    disc = []
    for it in os.listdir(d_path):
        fp = os.path.join(d_path, it)
        if os.path.isdir(fp):
            mt = datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%d/%m/%Y %H:%M")
            disc.append((it, mt, "Pasta", "--"))
    for d in disc:
        tree.insert("", tk.END, values=d)

    def confirmar_selecao_arquivos():
        s = tree.selection()
        if not s:
            messagebox.showwarning("Atenção", "Nenhuma disciplina selecionada.")
            return
        v = tree.item(s[0])["values"]
        disc_nome = v[0]
        p_disc = os.path.join(d_path, disc_nome)
        if not os.path.isdir(p_disc):
            messagebox.showerror("Erro", f"A pasta da disciplina '{p_disc}' não foi encontrada.")
            return

        expected_folder = "1.ENTREGAS"
        match_entrega = None

        def normalize(n):
            return n.lower().replace(" ", "").replace("-", "").replace("_", "").replace(".", "")

        for folder in os.listdir(p_disc):
            if normalize(folder) == normalize(expected_folder):
                match_entrega = folder
                break
        if not match_entrega:
            messagebox.showerror("Erro", f"A pasta de entrega '{expected_folder}' não foi encontrada.")
            return

        p_ent = os.path.join(p_disc, match_entrega)
        if not os.path.isdir(p_ent):
            messagebox.showerror("Erro", f"A pasta de entrega '{p_ent}' não pôde ser acessada.")
            return

        sel_arq = filedialog.askopenfilenames(
            title="Selecione arquivos para entrega",
            initialdir=p_ent
        )
        if not sel_arq:
            messagebox.showwarning("Atenção", "Nenhum arquivo foi selecionado.")
            return

        proc = []
        for arq in sel_arq:
            n_arq = os.path.basename(arq)
            d_ext = extrair_dados_arquivo(n_arq)
            d_ext["caminho"] = arq
            proc.append(d_ext)
        if not proc:
            messagebox.showerror("Erro", "Nenhum dado foi processado.")
            return

        discip_win.destroy()
        exibir_interface_tabela(
            numero,
            arquivos_previos=proc,
            caminho_projeto=caminho,
            pasta_entrega=p_ent,
            master=master
        )

    def voltar():
        discip_win.destroy()
        if master is not None:
            master.deiconify()

    cf = tk.Frame(discip_win, bg="#f5f5f5", padx=20, pady=20)
    cf.pack(fill=tk.BOTH, expand=True)
    bf = tk.Frame(discip_win)
    bf.pack(fill=tk.X, pady=5, padx=10)
    ttk.Button(bf, text="Voltar", command=voltar).pack(side=tk.LEFT, padx=5)
    ttk.Button(bf, text="Confirmar Seleção", command=confirmar_selecao_arquivos).pack(side=tk.RIGHT, padx=5)

def carregar_json(fp):
    if os.path.exists(fp):
        with open(fp, "r", encoding="utf-8") as f:
            return json.load(f)
    else:
        return {}

def salvar_json(fp, data):
    try:
        with open(fp, "w", encoding="utf-8") as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar dados em JSON: {e}")

def extrair_dados_arquivo(nome_arquivo):
    nb, ext = os.path.splitext(nome_arquivo)
    pt = nb.split('-')
    try:
        cr = pt[7] if len(pt)>7 else ""
        if '.' in cr:
            cs = cr.split('.')
            conj = cs[0]
            num_doc = cs[1]
        else:
            conj = cr
            num_doc = ""
        d = {
            "Status": pt[0] if len(pt)>0 else "",
            "Cliente": pt[1] if len(pt)>1 else "",
            "N° do Projeto": pt[2] if len(pt)>2 else "",
            "Organização": pt[3] if len(pt)>3 else "",
            "Sigla da Disciplina": pt[4] if len(pt)>4 else "",
            "Fase": pt[5] if len(pt)>5 else "",
            "Tipo de Documento": pt[6] if len(pt)>6 else "",
            "Conjunto": conj,
            "N° do Documento": num_doc,
            "Bloco": pt[8] if len(pt)>8 else "",
            "Pavimento": pt[9] if len(pt)>9 else "",
            "Subsistema": pt[10] if len(pt)>10 else "",
            "Tipo do Desenho": pt[11] if len(pt)>11 else "",
            "Revisão": pt[12] if len(pt)>12 else "",
            "Nome do Arquivo": nome_arquivo,
            "Extensão": ext.strip('-'),
            "Modificação": datetime.now().strftime("%d/%m/%Y"),
            "Modificado por": "Usuário"
        }
    except IndexError:
        d = {
            "Status": "",
            "Cliente": "",
            "N° do Projeto": "",
            "Organização": "",
            "Sigla da Disciplina": "",
            "Fase": "",
            "Tipo de Documento": "",
            "Conjunto": "",
            "N° do Documento": "",
            "Bloco": "",
            "Pavimento": "",
            "Subsistema": "",
            "Tipo do Desenho": "",
            "Revisão": "",
            "Nome do Arquivo": nome_arquivo,
            "Extensão": ext.strip('-'),
            "Modificação": datetime.now().strftime("%d/%m/%Y"),
            "Modificado por": "Usuário"
        }
    return d

def exibir_interface_tabela(
    numero: str,
    arquivos_previos: list[dict] | None = None,
    caminho_projeto: str | None = None,
    pasta_entrega: str | None = None,
    master=None
): 
    exibir_win = tk.Toplevel(master)
    exibir_win.title(f"Gerenciador de Projetos - Projeto {numero}")
    exibir_win.geometry("1200x800") 

    fpr = tk.Frame(exibir_win)
    fpr.pack(fill=tk.BOTH, expand=True)

    bar_l = tk.Frame(fpr, bg="#2c3e50", width=200)
    bar_l.pack(side=tk.LEFT, fill=tk.Y)
    lbl_t = tk.Label(bar_l, text="OAE - Engenharia", font=("Helvetica",14,"bold"),
                     bg="#2c3e50", fg="white")
    lbl_t.pack(pady=10)
    lbl_pj = tk.Label(bar_l, text="PROJETOS", font=("Helvetica",10,"bold"),
                      bg="#34495e", fg="white", anchor="w", padx=10)
    lbl_pj.pack(fill=tk.X, pady=5)
    ls_pj = tk.Listbox(bar_l, height=5, bg="#ecf0f1", font=("Helvetica",9))
    ls_pj.pack(fill=tk.X, padx=10, pady=5, anchor="center")
    ls_pj.insert(tk.END, f"Projeto {numero}")
    lbl_m = tk.Label(bar_l, text="MEMBROS", font=("Helvetica",10,"bold"),
                     bg="#34495e", fg="white", anchor="w", padx=10)
    lbl_m.pack(fill=tk.X, pady=5)
    bar_l.pack_propagate(False)

    cp = tk.Frame(fpr)
    cp.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

    def fazer_analise_nomenclatura():
        la = []
        for it in tabela.get_children():
            v = tabela.item(it)["values"]
            n_arq = v[1]
            cmin = v[10]
            dx = extrair_dados_arquivo(n_arq)
            dx["caminho"] = cmin
            la.append(dx)
        if not la:
            messagebox.showinfo("Aviso", "Nenhum arquivo adicionado para análise.")
        else:
            exibir_win.withdraw()
            tela_analise_nomenclatura(
                numero,              # passamos o número do projeto
                la,
                pasta_entrega=pasta_entrega,
                master=exibir_win
            )

    lbl_i = tk.Label(cp, text="Adicionar Arquivos para Entrega",
                     font=("Helvetica",15,"bold"), anchor="w")
    lbl_i.place(x=10, y=10)

    fb = tk.Frame(cp)
    fb.pack(side=tk.TOP, anchor="ne", pady=10, padx=10)
    ttk.Button(fb, text="Fazer análise da Nomenclatura",
               command=fazer_analise_nomenclatura).pack(side=tk.LEFT, padx=5)

    cols = ["Status","Nome do Arquivo","Extensão","Nº do Arquivo",
            "Fase","Tipo","Revisão","Modificação","Modificado por",
            "Entrega","caminho"]
    tabela = ttk.Treeview(cp, columns=cols, show="headings", height=20)
    for c in cols:
        tabela.heading(c, text=c)
        if c=="Nome do Arquivo":
            tabela.column(c, width=300)
        elif c=="caminho":
            tabela.column(c, width=0, stretch=False, minwidth=0)
        else:
            tabela.column(c, width=120)
    tabela["displaycolumns"] = (
        "Status","Nome do Arquivo","Extensão","Nº do Arquivo",
        "Fase","Tipo","Revisão","Modificação","Modificado por","Entrega"
    )
    tabela.pack(fill=tk.BOTH, expand=True)

    if arquivos_previos:
        for d_ext in arquivos_previos:
            tabela.insert("", tk.END, values=(
                d_ext.get("Status",""), d_ext.get("Nome do Arquivo",""), d_ext.get("Extensão",""),
                d_ext.get("N° do Arquivo",""), d_ext.get("Fase",""), d_ext.get("Tipo de Documento",""),
                d_ext.get("Revisão",""), d_ext.get("Modificação",""), d_ext.get("Modificado por",""),
                "",  # campo "Entrega" temporário
                d_ext.get("caminho","")
            ))

    def adicionar_arquivos():
        ar = filedialog.askopenfilenames(title="Selecione arquivos")
        for a in ar:
            n = os.path.basename(a)
            dx = extrair_dados_arquivo(n)
            dx["caminho"] = a
            tabela.insert("", tk.END, values=(
                dx.get("Status",""), dx.get("Nome do Arquivo",""), dx.get("Extensão",""),
                dx.get("N° do Documento",""), dx.get("Fase",""), dx.get("Tipo de Documento",""),
                dx.get("Revisão",""), dx.get("Modificação",""), dx.get("Modificado por",""),
                "",  # campo "Entrega"
                dx.get("caminho","")
            ))

    def remover_arquivo():
        s = tabela.selection()
        if s:
            for i in s:
                tabela.delete(i)
        else:
            messagebox.showinfo("Informação", "Nenhum item selecionado.")

    def voltar():
        exibir_win.destroy()
        if master is not None:
            master.deiconify()

    bf2 = tk.Frame(cp)
    bf2.pack(side=tk.LEFT, pady=10, padx=10)
    ttk.Button(bf2, text="Adicionar Arquivo", command=adicionar_arquivos).pack(side=tk.LEFT, padx=5)
    ttk.Button(bf2, text="Remover Arquivo", command=remover_arquivo).pack(side=tk.LEFT, padx=5)

    bf3 = tk.Frame(cp)
    bf3.pack(side="bottom", anchor="e", pady=5, padx=10)
    ttk.Button(bf3, text="Voltar", command=voltar).pack(side=tk.LEFT, padx=5)
    ttk.Button(bf3, text="Sair", command=exibir_win.destroy).pack(side=tk.RIGHT, padx=5)


def tela_analise_nomenclatura(projeto_num: str, lista_arquivos: list[dict], pasta_entrega: str, master=None):
    logging.debug(">>> INDO PARA tela_analise_nomenclatura: projeto=%s, pasta_entrega=%s, total_arquivos=%d", projeto_num, pasta_entrega, len(lista_arquivos))
    esquema = carregar_regras_nomenclatura(projeto_num)

    logging.debug("… regras de nomenclatura carregadas: %s", esquema.get("campos", []))

    token_win = tk.Toplevel(master)
    token_win.title("Verificação de Nomenclatura (Tokens)")
    token_win.geometry("1200x700")

    lbl_instr = tk.Label(
        token_win,
        text="Confira a nomenclatura quebrada em tokens. Linhas em VERMELHO têm tokens incorretos; AMARELO, tokens faltando.",
        font=("Helvetica",12)
    )
    lbl_instr.pack(pady=MARGIN_SIZE)

    frm_botoes = tk.Frame(token_win)
    frm_botoes.pack(fill=tk.X, padx=10, pady=5)

    # Botão “Mostrar Padrão” agora chama a versão modificada, que reconstrói o nome usando
    # o arquivo selecionado no Treeview e insere os separadores corretos.
    btn_mostrar = tk.Button(
        frm_botoes,
        text="Mostrar Padrão",
        command=lambda: mostrar_nomenclatura_padrao(esquema, lista_arquivos, tree, lista_tokens_por_arquivo, master=token_win)
    )
    btn_mostrar.pack(side=tk.LEFT, padx=5)

    btn_voltar = tk.Button(
        frm_botoes,
        text="Voltar",
        command=lambda: (_voltar_para_exibir(token_win, master))
    )
    btn_voltar.pack(side=tk.LEFT, padx=5)

    btn_avancar = tk.Button(
        frm_botoes,
        text="Avançar",
        command=lambda: (_tentar_avancar(token_win, esquema, lista_arquivos, pasta_entrega))
    )
    btn_avancar.pack(side=tk.RIGHT, padx=5)

    # --- monta lista de tokens para cada arquivo ---
    max_tokens = 0
    lista_tokens_por_arquivo: list[list[str]] = []
    for a in lista_arquivos:
        nome = a.get("Nome do Arquivo", "")
        nome_sem_ext, _ = os.path.splitext(nome)
        tokens = split_including_separators(nome_sem_ext, esquema)
        lista_tokens_por_arquivo.append(tokens)
        if len(tokens) > max_tokens:
            max_tokens = len(tokens)

    col_names = [f"T{i}" for i in range(1, max_tokens + 1)]
    frame_tv = tk.Frame(token_win)
    frame_tv.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

    sb_y = tk.Scrollbar(frame_tv, orient="vertical")
    sb_y.pack(side=tk.RIGHT, fill=tk.Y)
    tree = ttk.Treeview(
        frame_tv,
        columns=col_names,
        show="headings",
        yscrollcommand=sb_y.set
    )
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    sb_y.config(command=tree.yview)

    for cn in col_names:
        tree.heading(cn, text="")
        tree.column(cn, width=80, anchor="center")

    tree.tag_configure("ok",       background="white")
    tree.tag_configure("mismatch", background="#FF9999")
    tree.tag_configure("missing",  background="#FFFF99")

    for idx, a in enumerate(lista_arquivos):
        tokens = lista_tokens_por_arquivo[idx]
        tags_result = verificar_tokens(tokens, esquema)

        row_vals = []
        row_tags = []
        for i in range(max_tokens):
            if i < len(tokens):
                row_vals.append(tokens[i])
            else:
                row_vals.append("")

            if i < len(tags_result):
                if tags_result[i] == "mismatch":
                    row_tags.append("mismatch")
                elif tags_result[i] == "missing":
                    row_tags.append("missing")
                else:
                    row_tags.append("ok")
            else:
                row_tags.append("missing")

        if any(tg == "mismatch" for tg in row_tags):
            tag_linha = "mismatch"
        elif any(tg == "missing" for tg in row_tags):
            tag_linha = "missing"
        else:
            tag_linha = "ok"

        tree.insert("", tk.END, values=row_vals, tags=(tag_linha,))

    def mostrar_nomenclatura_padrao(nomenclatura_json: dict, lista_arquivos: list[dict], treeview: ttk.Treeview, lista_tokens_por_arquivo: list[list[str]], master=None):
        sel = treeview.selection()
        if not sel:
            messagebox.showinfo("Info", "Selecione uma linha para ver o padrão.")
            return

        # índice da linha selecionada
        idx = treeview.index(sel[0])

        # 1) obtém tokens completos (incluindo separadores) do arquivo
        tokens_com_seps = lista_tokens_por_arquivo[idx]

        # 2) extrai apenas os “valores” (filtrando todos os separadores '-' e '.')
        valores = [t for t in tokens_com_seps if t not in ['-', '.']]

        # 3) núm. de campos (sem contar revisão)
        campos_cfg = nomenclatura_json.get("campos", [])
        num_campos = len(campos_cfg)

        # 4) separa os valores que correspondem aos campos (0 até num_campos-1)
        #    e, se sobrar, considera o próximo como token de revisão
        valores_campos = valores[:num_campos]
        if len(valores) > num_campos:
            rev_token = valores[num_campos]
        else:
            # se não veio revisão no nome, criamos um valor padrão “1” ou “A…” igual ao preview
            rev_opcao    = nomenclatura_json.get("revisao_opcao", "Numérico")
            ndig         = nomenclatura_json.get("revisao_ndigitos", 2)
            prefixo      = nomenclatura_json.get("revisao_prefixo", "R")
            if rev_opcao == "Numérico":
                corpo = "1".rjust(ndig, "0")
            else:
                corpo = "A" * ndig
            rev_token = prefixo + corpo

        # 5) Reconstrói a string usando o valor de cada campo + separador correto
        partes = []
        for i, cinfo in enumerate(campos_cfg):
            # se não houver valor (campo ausente), deixamos em branco
            val_i = valores_campos[i] if i < len(valores_campos) else ""
            sep_i = cinfo.get("separador", "")
            partes.append(val_i)
            # só coloca separador se não for o último campo
            if i < num_campos - 1:
                partes.append(sep_i)

        # 6) acrescenta a parte de revisão
        sep_rev    = nomenclatura_json.get("revisao_separador", "-")
        prefixo    = nomenclatura_json.get("revisao_prefixo", "R")
        # caso o rev_token não tenha prefixo, anexamos
        if not rev_token.startswith(prefixo):
            rev_token = prefixo + rev_token
        partes.append(sep_rev + rev_token)

        nome_corrigido = "".join(partes)

        # 7) exibe numa janela para o usuário
        win = tk.Toplevel(master)
        win.title("Padrão Correto do Arquivo Selecionado")
        win.geometry("800x200")
        ttk.Label(win, text="Nome corrigido para este arquivo:", font=("Arial", 11, "bold")).pack(pady=10)
        txt = tk.Text(win, height=2, wrap="none", font=("Courier New", 12))
        txt.insert("1.0", nome_corrigido)
        txt.config(state="disabled")
        txt.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        ttk.Button(win, text="Fechar", command=win.destroy).pack(pady=5)

    def _voltar_para_exibir(token_window, exibir_window):
        token_window.destroy()
        exibir_window.deiconify()

    def _tentar_avancar(token_window, esquema_json, lista_arquivos_av, pasta_entrega):
        # validação de tokens
        for iid in tree.get_children():
            tags_ = tree.item(iid, "tags")
            if "mismatch" in tags_ or "missing" in tags_:
                messagebox.showwarning(
                    "Atenção",
                    "Há arquivos com tokens incorretos (VERMELHO) ou faltando (AMARELO).\n"
                    "Corrija o nome do arquivo (no sistema de arquivos) e volte para reanalisar."
                )
                return

        try:
            logging.debug(">>> _tentar_avancar: abrindo modal de tipo")
            # Em vez de ocultar:
            # token_window.withdraw()
            token_window.attributes('-disabled', True)

            def _on_tipo_escolhido(tipo):
                # reabilita a janela de tokens (caso ela precise ser mostrada depois)
                token_window.attributes('-disabled', False)
                token_window.withdraw()           # agora sim pode esconder
                tela_verificacao_revisao(
                    lista_arquivos_av,
                    pasta_entrega,
                    tipo,
                    master=token_window
                )

            # abre modal
            modal_tipo_entrega(token_window, on_confirm=_on_tipo_escolhido)

        except Exception:
            logging.exception("Falha em _tentar_avancar")
            messagebox.showerror("Erro", "Falha interna. Veja debug_entregas.log.")

def modal_tipo_entrega(master, on_confirm):
    logging.debug(">>> INDO PARA modal_tipo_entrega")
    win = tk.Toplevel(master)
    win.title("Tipo de Entrega")
    win.transient(master)
    win.grab_set()
    win.resizable(False, False)
    win.minsize(300, 140)
    master.update_idletasks()
    x = master.winfo_rootx() + master.winfo_width() // 2 - 150
    y = master.winfo_rooty()  + master.winfo_height() // 2 - 70
    win.geometry(f"+{x}+{y}")

    ttk.Label(win, text="Escolha o tipo de entrega:", font=("Arial", 11, "bold")).pack(
        pady=(12, 6), anchor="w", padx=20
    )
    var = tk.StringVar(value="AP")
    frm = ttk.Frame(win)
    frm.pack(padx=20, pady=5, anchor="w")
    ttk.Radiobutton(frm, text="Anteprojeto – 1.AP", value="AP", variable=var).pack(anchor="w")
    ttk.Radiobutton(frm, text="Projeto Executivo – 2.PE", value="PE", variable=var).pack(anchor="w")
    ttk.Separator(win).pack(fill="x", pady=10, padx=5)
    ttk.Button(
        win,
        text="Confirmar",
        command=lambda: (win.grab_release(), win.destroy(), on_confirm(var.get()))
    ).pack(pady=(0,12))


def tela_verificacao_revisao(lista_arquivos: list[dict], pasta_entrega: str, tipo: str, master=None):
    logging.debug(">>> INDO PARA tela_verificacao_revisao: tipo=%s, pasta_entrega=%s, num_arquivos=%d", tipo, pasta_entrega, len(lista_arquivos))

    arrv, aobs = identificar_revisoes(lista_arquivos)

    logging.debug("…arquivos revisados: %s | obsoletos: %s", [a["Nome do Arquivo"] for a in arrv], [a["Nome do Arquivo"] for a in aobs])
    
    rev_win = tk.Toplevel(master)
    rev_win.title("Verificação de Revisão")
    rev_win.geometry("1000x700")

    lb_i = tk.Label(rev_win, text="Confira os arquivos revisados e obsoletos antes da entrega.")
    lb_i.pack(pady=10)

    fr_r = tk.LabelFrame(rev_win, text="Arquivos Revisados")
    fr_r.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    tr_r = ttk.Treeview(fr_r, columns=["Nome do Arquivo","Revisão"], show="headings", height=10)
    tr_r.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    tr_r.heading("Nome do Arquivo", text="Nome do Arquivo")
    tr_r.heading("Revisão", text="Revisão")
    for a in arrv:
        tr_r.insert("", tk.END, values=(a["Nome do Arquivo"], a["Revisão"]))

    fr_o = tk.LabelFrame(rev_win, text="Arquivos Obsoletos")
    fr_o.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    tr_o = ttk.Treeview(fr_o, columns=["Nome do Arquivo","Revisão"], show="headings", height=10)
    tr_o.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    tr_o.heading("Nome do Arquivo", text="Nome do Arquivo")
    tr_o.heading("Revisão", text="Revisão")
    for a in aobs:
        tr_o.insert("", tk.END, values=(a["Nome do Arquivo"], a["Revisão"]))

    def voltar():
        rev_win.destroy()
        if master is not None:
            master.deiconify()

    def confirmar():
        try:
            caminhos = [Path(a["caminho"]) for a in (arrv + aobs)]
            pasta_raiz_entregas = Path(pasta_entrega)
            nova = processar_entrega_arquivos_tipo(caminhos, pasta_raiz_entregas, tipo)
            messagebox.showinfo(
                "Sucesso",
                f"Nova entrega criada:\n{nova}\n"
                "_controle_entrega.json gerado com o status dos arquivos."
            )
            rev_win.destroy()
            if master is not None:
                master.destroy()
            sys.exit(0)
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao processar entrega:\n{e}")

    bf = tk.Frame(rev_win)
    bf.pack(side="bottom", anchor="e", pady=5, padx=10)
    ttk.Button(bf, text="Voltar", command=voltar).pack(side=tk.LEFT, padx=5)
    ttk.Button(bf, text="Confirmar", command=confirmar).pack(side=tk.RIGHT, padx=5)
    ttk.Button(rev_win, text="Fechar", command=rev_win.destroy).pack(pady=10)

    rev_win.mainloop()

def _lista_erros_treeview(erros):
    w = tk.Toplevel()
    w.title("Lista de erros")
    tree = ttk.Treeview(w, columns=["e"], show="headings", height=10)
    tree.heading("e", text="Ocorrências")
    tree.pack(fill="both", expand=True, padx=10, pady=10)
    for e in erros:
        tree.insert("", "end", values=(e,))
    ttk.Button(w, text="Fechar", command=w.destroy).pack(pady=5)

def identificar_revisoes(lista_arquivos):
    grupos = {}
    for a in lista_arquivos:
        nb, _ = os.path.splitext(a["Nome do Arquivo"])
        t = nb.split("-")
        if len(t)<2:
            continue
        idf = "-".join(t[:-1])
        rev = t[-1] if t[-1].startswith("R") and t[-1][1:].isdigit() else "R00"
        grupos.setdefault(idf, []).append((rev, a))
    arrv = []
    aobs = []
    for idf, arqs in grupos.items():
        arqs.sort(key=lambda x: int(x[0][1:]) if x[0][1:].isdigit() else 0)
        rm = arqs[-1][1]
        arrv.append(rm)
        aobs.extend([q[1] for q in arqs[:-1]])
    return arrv, aobs

def _calc_md5(path: Path, buf=8192) -> str | None:
    if not path.exists():
        return None
    h = hashlib.md5()
    with path.open("rb") as f:
        while chunk := f.read(buf):
            h.update(chunk)
    return h.hexdigest()


def _carregar_status_anterior(pasta_entrega_atual: Path) -> dict[str, dict]:
    """
    Varre a entrega anterior (a subpasta imediatamente marcada -OBSOLETO).
    Retorna dict nome→{"hash":…, "rev": "R03"} para comparação de versões.
    """
    ant = None
    for sib in pasta_entrega_atual.parent.iterdir():
        if sib.is_dir() and sib.name.endswith("-OBSOLETO"):
            ant = sib
    if not ant:
        return {}
    res = {}
    for f in ant.iterdir():
        if f.is_file():
            nome = f.name
            hash_ = _calc_md5(f)
            rev   = nome.rsplit("-R", 1)[-1] if "-R" in nome else ""
            res[nome] = {"hash": hash_, "rev": rev}
    return res


def _status_arquivo(arquivo: Path, info_ant: dict) -> str:
    nome = arquivo.name
    hash_atual = _calc_md5(arquivo)
    rev_atual  = nome.rsplit("-R", 1)[-1] if "-R" in nome else ""

    ant = info_ant.get(nome)
    if ant is None or ant["hash"] is None:          # não existia mais
        return "novo"

    if hash_atual == ant["hash"]:
        return "igual"
    if rev_atual > ant["rev"]:
        return "revisado"
    return "mod_sem_rev"


def criar_arquivo_controle(pasta_raiz_entregas: str) -> None:
    """
    Gera/atualiza GRD.xlsx no layout matricial.
    Requer existir <pasta>/historico_entregas.json.
    """
    hist_json = Path(pasta_raiz_entregas) / "historico_entregas.json"
    if not hist_json.exists():
        logging.warning("historico_entregas.json inexistente em %s", pasta_raiz_entregas)
        return
    historico = json.loads(hist_json.read_text(encoding="utf-8"))
    if not historico:
        logging.info("Histórico vazio, GRD não gerado.")
        return

    # 1. carrega template
    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active

    col_inicio_ent = 3  # A=Grupo, B=Extens., C = primeira entrega
    while ws.cell(row=5, column=col_inicio_ent).value:
        # apaga o cabeçalho
        ws.cell(row=5, column=col_inicio_ent).value = None
        # apaga intervalo de dados (linhas 6-2000, ajuste conforme precisar)
        for row in ws.iter_rows(min_row=6, max_row=2000,
                                min_col=col_inicio_ent, max_col=col_inicio_ent):
            for cell in row:
                cell.value = None
                cell.fill  = None
        col_inicio_ent += 1

    # map cores
    fill_verde   = PatternFill("solid", fgColor="C6EFCE")
    fill_azul    = PatternFill("solid", fgColor="9BC2E6")
    fill_laranja = PatternFill("solid", fgColor="FFC000")

    # 2. descobrir próxima coluna livre
    col_inicio_ent = 3  # A=Grupo, B=Extens., C = 1ª entrega
    col_atual = col_inicio_ent
    while ws.cell(row=5, column=col_atual).value:   # linha 5 tem cabeçalhos de entrega no template
        col_atual += 1

    # 3. para cada entrega no histórico (na ordem)
    for ent in historico:
        tipo   = ent.get("tipo_entrega", "EX")
        etapa  = ent.get("etapa", 1)
        idx    = str(col_atual - col_inicio_ent + 1).zfill(2)   # NN
        cabec  = f"Z.{tipo}.ENT {idx} - ENTREGUE"
        ws.cell(row=5, column=col_atual, value=cabec)

        # copia largura & validação da coluna anterior (se houver)
        if col_atual > col_inicio_ent:
            src_col = get_column_letter(col_atual - 1)
            dst_col = get_column_letter(col_atual)
            ws.column_dimensions[dst_col].width = ws.column_dimensions[src_col].width
            for dv in ws.data_validations.dataValidation:
                if dv.ranges and src_col in str(dv.ranges):
                    new_dv = DataValidation(
                        type=dv.type, formula1=dv.formula1, allow_blank=dv.allow_blank
                    )
                    new_dv.add(f"{dst_col}6:{dst_col}2000")   # mesmo range aproximado
                    ws.add_data_validation(new_dv)

        # coleta info da entrega anterior para definir status/cor
        pasta_entrega = Path(ent["pasta_entrega"])
        info_ant = _carregar_status_anterior(pasta_entrega)

        # 3b. preencher linhas (a partir da linha 8 em diante, uma linha por arquivo)
        linha = 8
        for nome in ent["arquivos_entregues"]:
            arq = pasta_entrega / nome
            extens = arq.suffix.upper()  # ".PDF"
            status = _status_arquivo(arq, info_ant)
            cor = {"novo": fill_verde,
                   "revisado": fill_azul,
                   "mod_sem_rev": fill_laranja}.get(status)

            # Grupo em branco (col-A)
            ws.cell(row=linha, column=1, value="")

            # Extens.
            ws.cell(row=linha, column=2, value=extens)

            # Celula da entrega
            c = ws.cell(row=linha, column=col_atual, value=nome)
            if cor:
                c.fill = cor
            linha += 1

        col_atual += 1  # próxima entrega → próxima coluna

    # 4. atualiza “Gerado em”
    ws["B3"].value = datetime.now().strftime("%d/%m/%Y %H:%M")

    # 5. salva
    out_path = Path(pasta_raiz_entregas) / "GRD.xlsx"
    wb.save(out_path)
    logging.info("GRD.xlsx atualizado: %s", out_path)

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    janela_selecao_projeto(root)
    root.mainloop()
